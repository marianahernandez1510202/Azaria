#!/usr/bin/env python3
"""Genera diagrama de Gantt en Excel para el proyecto Azaria."""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Gantt - Azaria"

# ========== COLORES ==========
VERDE = PatternFill(start_color="27AE60", end_color="27AE60", fill_type="solid")       # Completado
AMARILLO = PatternFill(start_color="F1C40F", end_color="F1C40F", fill_type="solid")     # Pendiente
ROJO = PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid")         # No implementado
AZUL = PatternFill(start_color="3498DB", end_color="3498DB", fill_type="solid")          # Nuevo implementado
NARANJA = PatternFill(start_color="E67E22", end_color="E67E22", fill_type="solid")       # Modificado
HEADER_FILL = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")   # Header
LEGEND_BG = PatternFill(start_color="ECF0F1", end_color="ECF0F1", fill_type="solid")     # Leyenda
WHITE_FILL = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
ROW_ALT = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")       # Filas alternas

WHITE_FONT = Font(color="FFFFFF", bold=True, size=11)
HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
BOLD_FONT = Font(bold=True, size=11)
NORMAL_FONT = Font(size=10)
CATEGORY_FONT = Font(bold=True, size=11, color="2C3E50")
SMALL_FONT = Font(size=9, color="FFFFFF", bold=True)

THIN_BORDER = Border(
    left=Side(style='thin', color='D5D8DC'),
    right=Side(style='thin', color='D5D8DC'),
    top=Side(style='thin', color='D5D8DC'),
    bottom=Side(style='thin', color='D5D8DC')
)

CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT = Alignment(horizontal='left', vertical='center', wrap_text=True)

# ========== FECHAS ==========
fecha_inicio = datetime(2025, 1, 12)
fecha_fin = datetime(2025, 3, 13)

# Generar semanas
semanas = []
current = fecha_inicio
while current <= fecha_fin:
    week_end = current + timedelta(days=6)
    if week_end > fecha_fin:
        week_end = fecha_fin
    semanas.append((current, week_end))
    current = week_end + timedelta(days=1)

# ========== DATOS DEL PROYECTO ==========
# (categoria, tarea, estado, semana_inicio, semana_fin, color)
# estado: "completado", "nuevo", "modificado", "pendiente", "no_implementado"
# semanas van del 0 al N

tareas = [
    # --- INFRAESTRUCTURA ---
    ("INFRAESTRUCTURA", "Configuracion inicial del proyecto", "completado", 0, 1),
    ("", "Base de datos (70+ tablas)", "completado", 0, 2),
    ("", "Backend MVC (PHP 8)", "completado", 0, 3),
    ("", "Frontend React (CRA 5)", "completado", 0, 3),
    ("", "PWA / Service Worker", "completado", 2, 4),
    ("", "Sistema de autenticacion (HMAC-SHA256)", "completado", 1, 3),

    # --- MODULOS CORE ---
    ("MODULOS CORE", "Login / Auth Flow", "completado", 1, 3),
    ("", "Dashboard Paciente", "completado", 2, 4),
    ("", "Dashboard Especialista", "completado", 3, 5),
    ("", "Dashboard Administrador", "completado", 3, 5),
    ("", "Perfil de usuario", "completado", 3, 4),

    # --- MODULOS MEDICOS ---
    ("MODULOS MEDICOS", "Nutricion (registro comidas, agua, checklist)", "completado", 3, 5),
    ("", "Planes Nutricionales (PDF, asignacion)", "completado", 5, 6),
    ("", "Medicina (glucosa, presion, dolor)", "completado", 3, 5),
    ("", "Fisioterapia (videos, rutinas, checklist)", "completado", 4, 6),
    ("", "Neuropsicologia (estados animo, ejercicios)", "modificado", 4, 7),
    ("", "Neuropsicologia - ACT/AAQ-2 Tools", "nuevo", 6, 7),
    ("", "Ortesis / Protesis (educativo, dispositivo)", "modificado", 4, 7),

    # --- COMUNICACION ---
    ("COMUNICACION", "Citas (agendar, cancelar, reagendar)", "modificado", 4, 7),
    ("", "Chat / Mensajes", "completado", 4, 6),
    ("", "Recordatorios", "completado", 5, 6),
    ("", "Outlook Calendar Sync", "completado", 5, 7),

    # --- CONTENIDO ---
    ("CONTENIDO", "Blog (articulos, likes, comentarios)", "modificado", 5, 7),
    ("", "Comunidad (publicaciones, reacciones)", "modificado", 5, 7),
    ("", "FAQs", "completado", 5, 6),

    # --- FUNCIONALIDADES NUEVAS ---
    ("NUEVAS FUNCIONALIDADES", "Expediente Clinico (resumen, archivos, compartir)", "nuevo", 6, 8),
    ("", "Expediente Compartido (link temporal 72h)", "nuevo", 7, 8),
    ("", "Configuracion (notificaciones, privacidad, seguridad)", "nuevo", 7, 8),
    ("", "ModuleLayout (navegacion inferior + boton volver)", "nuevo", 6, 7),
    ("", "Panel de Accesibilidad", "completado", 5, 6),
    ("", "Narrador de Voz (VoiceHelper)", "completado", 5, 6),

    # --- AUDITORIA Y CORRECCIONES ---
    ("AUDITORIA FEB 2025", "Auditoria completa del sistema", "nuevo", 8, 8),
    ("", "Correccion nombres de tablas en Models", "nuevo", 8, 8),
    ("", "Correccion ruta ortesis/checklist", "nuevo", 8, 8),
    ("", "Eliminacion rutas de test en produccion", "nuevo", 8, 8),
    ("", "Mover SQL inline a Controllers", "nuevo", 8, 8),
    ("", "Agregar 13 rutas faltantes", "nuevo", 8, 8),
    ("", "Limpieza error_log() sensibles", "nuevo", 8, 8),
    ("", "Correccion DatabaseService (seguridad)", "nuevo", 8, 8),

    # --- PENDIENTES ---
    ("PENDIENTES", "Modulo Fases - Backend (Controller + Model + Rutas)", "pendiente", 8, 8),
    ("", "Modulo Fases - Frontend (pagina + CSS)", "no_implementado", 8, 9),
    ("", "EmailService (todos los metodos son TODO)", "no_implementado", -1, -1),
    ("", "CSRF Protection", "no_implementado", -1, -1),
    ("", "Middleware sin usar (RateLimit, CORS, Role, Moderation)", "no_implementado", -1, -1),
    ("", "Tests (backend + frontend)", "no_implementado", -1, -1),
    ("", "Sistema de migraciones versionado", "no_implementado", -1, -1),
]

COLOR_MAP = {
    "completado": VERDE,
    "nuevo": AZUL,
    "modificado": NARANJA,
    "pendiente": AMARILLO,
    "no_implementado": ROJO,
}

# ========== LEYENDA (filas 1-3) ==========
ws.merge_cells('A1:C1')
ws['A1'] = 'AZARIA - Diagrama de Gantt del Proyecto'
ws['A1'].font = Font(bold=True, size=16, color="2C3E50")
ws['A1'].alignment = LEFT

# Leyenda en fila 2
leyenda_items = [
    ("Completado", VERDE),
    ("Pendiente", AMARILLO),
    ("No implementado", ROJO),
    ("Nuevo implementado", AZUL),
    ("Modificado", NARANJA),
]

col_leyenda = 1
ws.merge_cells('A2:C2')
ws['A2'] = 'Leyenda:'
ws['A2'].font = Font(bold=True, size=10, color="7F8C8D")

for i, (texto, color) in enumerate(leyenda_items):
    col = 4 + i * 2
    cell_color = ws.cell(row=2, column=col)
    cell_color.fill = color
    cell_color.border = THIN_BORDER
    cell_text = ws.cell(row=2, column=col + 1)
    cell_text.value = texto
    cell_text.font = Font(size=9, color="2C3E50")
    cell_text.alignment = LEFT

# ========== HEADER ROW (fila 4) ==========
header_row = 4
headers = ['#', 'Categoria', 'Tarea', 'Estado']
for i, h in enumerate(headers, 1):
    cell = ws.cell(row=header_row, column=i)
    cell.value = h
    cell.fill = HEADER_FILL
    cell.font = WHITE_FONT
    cell.alignment = CENTER
    cell.border = THIN_BORDER

# Semanas como headers
for i, (start, end) in enumerate(semanas):
    col = 5 + i
    cell = ws.cell(row=header_row, column=col)
    cell.value = f"{start.strftime('%d/%m')}\n{end.strftime('%d/%m')}"
    cell.fill = HEADER_FILL
    cell.font = Font(color="FFFFFF", bold=True, size=8)
    cell.alignment = CENTER
    cell.border = THIN_BORDER

# Sub-header con numeros de semana
sub_row = header_row  # Keep it on same row for compactness

# ========== DATOS (desde fila 5) ==========
data_start = 5
task_num = 0

for idx, (categoria, tarea, estado, sem_inicio, sem_fin) in enumerate(tareas):
    row = data_start + idx
    is_alt = idx % 2 == 1
    bg = ROW_ALT if is_alt else WHITE_FILL

    # Numero
    if categoria == "":
        task_num += 1
        cell_num = ws.cell(row=row, column=1)
        cell_num.value = task_num
        cell_num.font = NORMAL_FONT
        cell_num.alignment = CENTER
        cell_num.fill = bg
        cell_num.border = THIN_BORDER
    else:
        task_num += 1
        cell_num = ws.cell(row=row, column=1)
        cell_num.value = task_num
        cell_num.font = BOLD_FONT
        cell_num.alignment = CENTER
        cell_num.fill = bg
        cell_num.border = THIN_BORDER

    # Categoria
    cell_cat = ws.cell(row=row, column=2)
    cell_cat.value = categoria
    cell_cat.font = CATEGORY_FONT if categoria else NORMAL_FONT
    cell_cat.alignment = LEFT
    cell_cat.fill = bg
    cell_cat.border = THIN_BORDER

    # Tarea
    cell_task = ws.cell(row=row, column=3)
    cell_task.value = tarea
    cell_task.font = BOLD_FONT if categoria else NORMAL_FONT
    cell_task.alignment = LEFT
    cell_task.fill = bg
    cell_task.border = THIN_BORDER

    # Estado
    estado_labels = {
        "completado": "Completado",
        "nuevo": "Nuevo",
        "modificado": "Modificado",
        "pendiente": "Pendiente",
        "no_implementado": "Sin implementar",
    }
    cell_estado = ws.cell(row=row, column=4)
    cell_estado.value = estado_labels.get(estado, estado)
    cell_estado.font = Font(size=9, bold=True, color="FFFFFF")
    cell_estado.fill = COLOR_MAP[estado]
    cell_estado.alignment = CENTER
    cell_estado.border = THIN_BORDER

    # Barras de Gantt
    for s in range(len(semanas)):
        col = 5 + s
        cell = ws.cell(row=row, column=col)
        cell.border = THIN_BORDER
        cell.alignment = CENTER

        if sem_inicio >= 0 and sem_fin >= 0 and sem_inicio <= s <= sem_fin:
            cell.fill = COLOR_MAP[estado]
            # Poner texto solo en la primera celda de la barra
            if s == sem_inicio:
                cell.font = SMALL_FONT
        else:
            cell.fill = bg

# ========== RESUMEN AL FINAL ==========
summary_row = data_start + len(tareas) + 2
ws.merge_cells(f'A{summary_row}:C{summary_row}')
ws.cell(row=summary_row, column=1).value = 'RESUMEN DEL PROYECTO'
ws.cell(row=summary_row, column=1).font = Font(bold=True, size=14, color="2C3E50")

counts = {"completado": 0, "nuevo": 0, "modificado": 0, "pendiente": 0, "no_implementado": 0}
for _, _, estado, _, _ in tareas:
    counts[estado] += 1

total = sum(counts.values())
summary_items = [
    (f"Modulos completados: {counts['completado']}", VERDE),
    (f"Nuevos implementados: {counts['nuevo']}", AZUL),
    (f"Modificados/Mejorados: {counts['modificado']}", NARANJA),
    (f"Pendientes: {counts['pendiente']}", AMARILLO),
    (f"Sin implementar: {counts['no_implementado']}", ROJO),
]

for i, (texto, color) in enumerate(summary_items):
    r = summary_row + 1 + i
    cell_dot = ws.cell(row=r, column=1)
    cell_dot.fill = color
    cell_dot.border = THIN_BORDER
    ws.merge_cells(f'B{r}:D{r}')
    cell_text = ws.cell(row=r, column=2)
    cell_text.value = texto
    cell_text.font = Font(size=11, bold=True)
    cell_text.alignment = LEFT

# Progreso total
prog_row = summary_row + len(summary_items) + 2
ws.merge_cells(f'A{prog_row}:D{prog_row}')
implementado = counts['completado'] + counts['nuevo'] + counts['modificado']
porcentaje = round(implementado / total * 100, 1)
ws.cell(row=prog_row, column=1).value = f'Progreso total: {implementado}/{total} tareas ({porcentaje}%)'
ws.cell(row=prog_row, column=1).font = Font(bold=True, size=13, color="27AE60")

# ========== ANCHOS DE COLUMNA ==========
ws.column_dimensions['A'].width = 4
ws.column_dimensions['B'].width = 22
ws.column_dimensions['C'].width = 48
ws.column_dimensions['D'].width = 15
for i in range(len(semanas)):
    ws.column_dimensions[get_column_letter(5 + i)].width = 12

# ========== FIJAR PANELES ==========
ws.freeze_panes = 'E5'

# ========== ROW HEIGHTS ==========
ws.row_dimensions[1].height = 30
ws.row_dimensions[2].height = 20
ws.row_dimensions[header_row].height = 35
for i in range(len(tareas)):
    ws.row_dimensions[data_start + i].height = 28

# ========== GUARDAR ==========
output_path = r"c:\Users\maria\Escritorio\Azaria\Gantt_Azaria.xlsx"
wb.save(output_path)
print(f"Gantt guardado en: {output_path}")
print(f"Total tareas: {total}")
print(f"Completadas: {counts['completado']}")
print(f"Nuevas: {counts['nuevo']}")
print(f"Modificadas: {counts['modificado']}")
print(f"Pendientes: {counts['pendiente']}")
print(f"Sin implementar: {counts['no_implementado']}")
print(f"Progreso: {porcentaje}%")
